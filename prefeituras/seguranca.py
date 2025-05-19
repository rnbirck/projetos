# %%
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine
import time

start_time = time.time()

caminho_destino = (
    "D:/OneDrive - Associacao Antonio Vieira/UAPP_ProjetoCEI/Prefeituras/Segurança"
)

# Configuração do banco de dados
usuario = "root"
senha = "ceiunisinos"
host = "localhost"
banco = "cei"

# Criar a conexão com MySQL
engine = create_engine(f"mysql+pymysql://{usuario}:{senha}@{host}/{banco}")
query_municipio = "SELECT id_municipio, municipio FROM municipio WHERE sigla_uf = 'RS'"

trad_mun = pd.read_sql_query(query_municipio, engine).assign(
    municipio_seg=lambda x: x["municipio"]
    .str.upper()
    .str.normalize("NFKD")
    .str.encode("ascii", errors="ignore")
    .str.decode("utf-8")
    .str.replace("-", "", regex=False)
    .str.replace(" ", "", regex=False)
    .str.replace("'", "", regex=False)
)


# FUNCOES
def leitura_indicadores_criminais(ano: int):
    """
    Lê os indicadores criminais de um determinado ano e retorna um DataFrame.
    """
    # Lê o arquivo Excel
    file_path = f"data/seguranca/indicadores_criminais_{ano}.xlsx"
    workbook = load_workbook(file_path, read_only=True)
    sheets = workbook.sheetnames

    def leitura_planilhas(sheet):
        # Le as planilhas do arquivo
        df = pd.read_excel(file_path, sheet_name=sheet, engine="calamine")
        df = df.assign(
            Ano=ano,
            Mês=sheet,
            Município=lambda x: x["Município"]
            .str.upper()
            .str.normalize("NFKD")
            .str.encode("ascii", errors="ignore")
            .str.decode("utf-8")
            .str.replace("-", "", regex=False)
            .str.replace(" ", "", regex=False)
            .str.replace("'", "", regex=False)
            .str.replace("DRMAURICIOCARDOSO", "DOUTORMAURICIOCARDOSO", regex=False),
        )
        return df

    return pd.concat([leitura_planilhas(sheet) for sheet in sheets], ignore_index=True)


def leitura_violencia_mulher(ano: int):
    """
    Lê os indicadores de violência contra a mulher de um determinado ano e retorna um DataFrame.
    """
    # Lê o arquivo Excel
    file_path = f"data/seguranca/violencia_mulher_{ano}.xlsx"
    workbook = load_workbook(file_path, read_only=True)

    def leitura_planilhas(sheet):
        df = pd.read_excel(file_path, sheet_name=sheet, engine="calamine")
        df = df.assign(
            Ano=ano,
            Indicador=sheet,
            Município=lambda x: x["Município"]
            .str.upper()
            .str.normalize("NFKD")
            .str.encode("ascii", errors="ignore")
            .str.decode("utf-8")
            .str.replace("-", "", regex=False)
            .str.replace(" ", "", regex=False)
            .str.replace("'", "", regex=False)
            .str.replace("DRMAURICIOCARDOSO", "DOUTORMAURICIOCARDOSO", regex=False),
        )
        return df

    return pd.concat(
        [leitura_planilhas(sheet) for sheet in workbook.sheetnames], ignore_index=True
    )


def ajustar_df_violencia_mulher(df):
    """
    Ajusta o df de violência contra a mulher.
    """
    df = (
        df.melt(
            id_vars=["Município", "Ano", "Indicador"],
            var_name="Mês",
            value_name="Valor",
        )
        .pivot(index=["Município", "Ano", "Mês"], columns="Indicador", values="Valor")
        .reset_index()
    )
    return df


def processar_dados_seguranca_mensal(
    df_indicadores_criminais, df_violencia_mulher, trad_mun
):
    """Junta e processa os dados de segurança e mulher"""
    df = (
        df_indicadores_criminais.merge(
            df_violencia_mulher, on=["Município", "Ano", "Mês"], how="left"
        )
        .merge(trad_mun, left_on="Município", right_on="municipio_seg", how="left")
        .dropna(subset=["municipio_seg"])
        .drop(columns=["municipio_seg", "Município"])
        .pipe(
            lambda df: df[
                ["id_municipio", "municipio", "Ano", "Mês"]
                + [
                    col
                    for col in df.columns
                    if col not in ["id_municipio", "municipio", "Ano", "Mês"]
                ]
            ]
        )
    )
    return df


def processar_dados_mensal_furtos(ano: int, trad_mun):
    """
    Processa os dados de furtos mensais de um determinado ano e retorna um DataFrame.
    """
    file_name = f"data/seguranca/furtos_{ano}.csv"
    df = (
        pd.read_csv(
            file_name,
            sep=";",
            encoding="latin1",
            usecols=["Sequência", "Data Fato", "Municipio Fato", "Tipo Enquadramento"],
            dtype={"Sequência": str},
            engine="pyarrow",
        )
        .dropna(subset=["Municipio Fato"])
        .assign(
            **{
                "Data Fato": lambda x: pd.to_datetime(
                    x["Data Fato"], format="%d/%m/%Y"
                ),
                "ano": lambda x: x["Data Fato"].dt.year,
                "mes": lambda x: x["Data Fato"].dt.month,
                "Municipio Fato": lambda x: x["Municipio Fato"]
                .str.upper()
                .str.normalize("NFKD")
                .str.encode("ascii", errors="ignore")
                .str.decode("utf-8")
                .str.replace("-", "", regex=False)
                .str.replace(" ", "", regex=False),
            }
        )
        .query("`Tipo Enquadramento`.str.contains('FURTO', na=False)")
        .groupby(by=["ano", "mes", "Municipio Fato", "Tipo Enquadramento"])
        .agg(n_furtos=("Sequência", "nunique"))
        .reset_index()
        .merge(trad_mun, left_on="Municipio Fato", right_on="municipio_seg", how="left")
        .drop(columns=["municipio_seg", "Municipio Fato"])
        .rename(columns={"Tipo Enquadramento": "tipo_furtos"})
    )
    return df


# DataFrame com os indicadores criminais
anos = range(2019, 2026)
df_indicadores_criminais = pd.concat(
    [leitura_indicadores_criminais(ano=ano) for ano in anos], ignore_index=True
)

# DataFrame com os indicadores de violência contra a mulher
df_violencia_mulher = pd.concat(
    [leitura_violencia_mulher(ano) for ano in anos],
    ignore_index=True,
).pipe(lambda df: ajustar_df_violencia_mulher(df))
# Merge dos DataFrames
df_seguranca_mensal = processar_dados_seguranca_mensal(
    df_indicadores_criminais=df_indicadores_criminais,
    df_violencia_mulher=df_violencia_mulher,
    trad_mun=trad_mun,
).to_csv(caminho_destino + "/seguranca.csv", index=False)
print("Arquivo de seguranca mensal Salvo")

# Microdados Furtos
anos_furtos = range(2022, 2026)
df_furtos_mensal = pd.concat(
    [processar_dados_mensal_furtos(ano=ano, trad_mun=trad_mun) for ano in anos_furtos],
    ignore_index=True,
).to_csv(caminho_destino + "/furtos.csv", index=False)
end_time = time.time()
print("Arquivo de furtos mensal Salvo")
print(f"Tempo total de execução: {end_time - start_time:.2f} segundos")
